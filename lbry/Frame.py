import discord
from openpyxl import load_workbook

# load_wb = load_workbook("Data/Ship_Database.xlsx", data_only=True)
# load_RSI = load_wb['RSI']

class ShipFrame:
    def __init__(self, title, footer, mft, number, image):
        self._title = title
        self._footer = footer
        self._MFT = mft
        self._num = number
        self._image = image
    
    def shipInfo(self):

        embed = discord.Embed(
            title=(self._MFT.cell(self._num, 2).value), 
            description=(self._MFT.cell(self._num, 12).value), 
            colour = 00000000)
        embed.set_author(name= self._title)
        embed.add_field(name= "제조사", value=(self._MFT.cell(self._num, 3).value), inline= True)
        embed.add_field(name= "구현단계", value=(self._MFT.cell(self._num, 4).value), inline= False)
        embed.add_field(name= "전장/전폭/전고", value=(self._MFT.cell(self._num, 5).value), inline= True)
        embed.add_field(name= "최소/최대 승무원", value=(self._MFT.cell(self._num, 6).value), inline= True)
        embed.add_field(name= "순항/최대 속도", value=(self._MFT.cell(self._num, 7).value), inline= True)
        embed.add_field(name= "분류", value=(self._MFT.cell(self._num, 8).value), inline= True)
        embed.add_field(name= "출고중량", value=(self._MFT.cell(self._num, 9).value), inline= True)
        embed.add_field(name= "화물용적", value=(self._MFT.cell(self._num, 10).value), inline= True)
        embed.add_field(name= "가격($)", value=(self._MFT.cell(self._num, 11).value), inline= True)
        embed.set_image(url= self._image)
        embed.set_footer(text= self._footer)
        return embed
